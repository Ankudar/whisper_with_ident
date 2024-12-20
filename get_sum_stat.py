import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from calendar import monthrange

print("start sum stat")
# Путь к директории с файлами Excel
directory = '.\\personal_statistic\\base\\2024\\'
year = int(directory.split('\\')[-2])  # извлечение года из пути к директории

# Получение списка файлов Excel
files = [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f)) and f.endswith('.xlsx')]

data = []
summary_df = pd.DataFrame(data)

# Запись итогового DataFrame в новый файл Excel
summary_df.to_excel(f'summary_{year}.xlsx', index=False)

book = load_workbook(f'summary_{year}.xlsx')
sheet = book.active

# Запись имен файлов в первую строку
sheet.insert_rows(1)
sheet['A1'] = 'Дата/Имя'
for i, file in enumerate(files, start=2):
    cell = sheet.cell(row=1, column=i)
    cell.value = file
    cell.alignment = Alignment(text_rotation=90)  # Установка текста вертикальным

# Установка текста в ячейке A1 вертикальным
sheet['A1'].alignment = Alignment(text_rotation=90)

for column_cells in sheet.columns:
    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = 6

# Сохранение изменений
book.save(f'summary_{year}.xlsx')

# Загрузить файл 'summary.xlsx'
summary_book = load_workbook(f'summary_{year}.xlsx')
summary_sheet = summary_book.active

# Получить вкладки '01' до '12'
months = [str(i).zfill(2) for i in range(1, 13)]

# Обработать каждый файл в директории
for file in files:
    # Загрузить текущий файл
    current_book = load_workbook(f'{directory}{file}')

    for month in months:
        # Если вкладки не существует, пропустить ее
        if month not in current_book.sheetnames:
            continue

        current_sheet = current_book[month]

        # Найти столбец для текущего файла в 'summary.xlsx'
        try:
            current_column = [cell.value for cell in summary_sheet[1]].index(file) + 1
        except ValueError:  # Если столбец не найден, добавить его
            current_column = summary_sheet.max_column + 1
            summary_sheet.cell(row=1, column=current_column, value=file)

        # Создать ячейки для всех дней месяца
        days_in_month = monthrange(year, int(month))[1]
        for day in range(1, days_in_month + 1):
            date_str = f"{str(day).zfill(2)}.{month}."
            try:
                date_row = [str(c.value) for c in summary_sheet['A']].index(date_str) + 1
            except ValueError:  # Если дата не найдена, добавить ее
                date_row = summary_sheet.max_row + 1
                summary_sheet.cell(row=date_row, column=1, value=date_str)

        # Пройти по каждой ячейке в первой строке текущего файла
        for i, cell in enumerate(current_sheet[1], start=1):
            if cell.value:  # Если ячейка содержит дату
                # Проверить количество непустых ячеек в столбце
                non_empty_cells = sum(1 for _ in filter(None, (c.value for c in current_sheet[get_column_letter(i)])))
                if non_empty_cells <= 5:
                    continue  # Если в столбце 5 или менее заполненных ячеек, пропустить этот столбец

                # Преобразовать дату в строку
                date_str = str(cell.value)

                # Найти соответствующую строку для этой даты в 'summary.xlsx'
                try:
                    date_row = [str(c.value) for c in summary_sheet['A']].index(date_str) + 1
                except ValueError:  # Если дата не найдена, пропустить ее
                    continue

                # Получить значение для этой даты из текущего файла
                value = current_sheet.cell(row=2, column=i).value

                # Записать значение в соответствующую ячейку 'summary.xlsx'
                summary_sheet.cell(row=date_row, column=current_column, value=value)

# Сохранить изменения в файле 'summary.xlsx'
summary_book.save(f'summary_{year}.xlsx')
print("finish sum stat")
